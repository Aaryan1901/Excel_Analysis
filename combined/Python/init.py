import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.chart import DoughnutChart, Reference, DataLabelList
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# File paths
input_path = r"C:\Users\Aaryan\Documents\combined\combined.xlsx"
output_dir = r"C:\Users\Aaryan\Documents\combined\output"
output_path = os.path.join(output_dir, "summary_output.xlsx")

os.makedirs(output_dir, exist_ok=True)

# Read data
df = pd.read_excel(input_path, usecols=["series_name", "delivery_date"])
df["delivered"] = df["delivery_date"].notna()

# Program metrics
total_programs = len(df)
delivered_programs = df["delivered"].sum()
remaining_programs = total_programs - delivered_programs

# Channel metrics
channels_total = df["series_name"].nunique()
channel_stats = df.groupby("series_name")["delivered"].agg(['sum', 'count'])
channel_stats["ratio"] = channel_stats["sum"] / channel_stats["count"]

channels_completed = (channel_stats["ratio"] == 1.0).sum()
channels_pending = channels_total - channels_completed
near_completion = (channel_stats["ratio"] >= 0.9).sum()
focus_less_90 = (channel_stats["ratio"] < 0.9).sum()
top_focus_channels = channel_stats[channel_stats["ratio"] < 0.9].sort_values("ratio", ascending=False).head(20)

# Summary data
summary_data = [
    ["Programs", "Total", total_programs],
    ["Programs", "Delivered", delivered_programs],
    ["Programs", "Remaining", remaining_programs],
    ["Channels", "Total", channels_total],
    ["Channels", "Completed", channels_completed],
    ["Channels", "Pending", channels_pending],
    ["Focus", "Near-Completion (≥90 %)", near_completion],
    ["Focus", "Focus (<90 %)", focus_less_90],
    ["Focus", "Next Top Channels to focus", 20]
]

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Summary"

# Styling
header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center = Alignment(horizontal="center")

# Headers
headers = ["Section", "Metric", "Value"]
for col, head in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=head)
    cell.fill = header_fill
    cell.font = Font(bold=True)
    cell.alignment = center
    cell.border = border

# Data rows
for r, row in enumerate(summary_data, 2):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

# Add chart data to side cells
ws["E2"] = "Delivered"
ws["E3"] = "Remaining"
ws["F2"] = delivered_programs
ws["F3"] = remaining_programs

# Doughnut chart
chart = DoughnutChart()
chart.title = f"{(delivered_programs / total_programs):.1%} Delivered"
chart.dataLabels = DataLabelList()
chart.dataLabels.showPercent = True

labels = Reference(ws, min_col=5, min_row=2, max_row=3)
values = Reference(ws, min_col=6, min_row=2, max_row=3)
chart.add_data(values, titles_from_data=False)
chart.set_categories(labels)

# Set segment colors
chart.series[0].points[0].graphicalProperties.solidFill = "FF4C4C"  # Red (Delivered)
chart.series[0].points[1].graphicalProperties.solidFill = "4F81BD"  # Blue (Remaining)

ws.add_chart(chart, "H2")

# Add Top Focus Channels sheet
focus_ws = wb.create_sheet("TopFocusChannels")
focus_ws.append(["Channel (series_name)", "Delivery %"])
for name, row in top_focus_channels.iterrows():
    focus_ws.append([name, round(row["ratio"] * 100, 2)])

# Save the file
wb.save(output_path)
print(f"✅ Output saved to: {output_path}")
